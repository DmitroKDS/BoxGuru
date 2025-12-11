from aiogram import types
from aiogram.fsm.context import FSMContext
from . import inline_keyboard
from config import Config

import openpyxl
import io

import bot.handlers.start.start
import logging



async def ask(message: types.Message, state: FSMContext, user_id: int):

    lang = await bot.handlers.start.start.set_lang(user_id, message)
    if lang == None:
        return
    answ = len(Config.answers.get(user_id, []))
    if answ < len(Config.questions[lang]) and Config.questions[lang][answ]['text'] in ["36.В какой стране ты проживаешь?", "36. У якій країні ти проживаєш?"]:
        Config.answers.setdefault(user_id, []).append("Россія или Украина")
        answ+=1

    if answ < len(Config.questions[lang]):
        options = [(option, f"answ_{i}") for i, option in enumerate(Config.questions[lang][answ]["options"])]+[(Config.translates['Пройти опрос еще раз'][lang], 'again')]
        await message.answer(
            Config.questions[lang][answ]['text'],
            reply_markup=inline_keyboard.create(*options), 
            disable_web_page_preview=True
        )
    else:
        await message.answer(
Config.translates[f"""Надеюсь, что ты не торопился и внимательно отвечал на мои вопросы. И если это так, тогда у меня получится подобрать для тебя просто ИДЕАЛЬНЫЙ смартфон в рамках твоего бюджета.
Я не торгую смартфонами и не стараюсь впарить тебе что-то, на чем я заработаю. Для меня главное, чтобы ты остался доволен и получал удовольствие, когда будешь пользоваться новым смартфоном.

Стоимость подбора смартфона 100 звёзд – всего 2$. Думаю, ты согласишься, что это небольшая сумма за подбор идеального смартфона.
«😩Не получается купить звёзды? Введи в YouTube запрос: как купить звёзды в Telegram?» И получишь четкую инструкцию🤩"""][lang],
            reply_markup=inline_keyboard.create((Config.translates['Пройти опрос еще раз'][lang], 'again')), 
            disable_web_page_preview=True
        )
        price = Config.pay_lang.get(user_id, "undefined")
        price = Config.prices.get("undefined")if price== None else Config.prices.get(price)

        await message.answer_invoice(
            title=Config.translates['Оплати для подбора смартфона'][lang],
            description=Config.translates["Оплата звездами"][lang],
            payload='private',
            prices=[
                types.LabeledPrice(label="XTR", amount=price),
            ],
            currency="XTR"
        )
        from bot.handlers.answer.payment import payment_test
        # await payment_test(message, state)





async def upload(buffer: io.BytesIO):
    buffer.seek(0)
    wb = openpyxl.load_workbook(buffer, data_only=True)
    questions = {}

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        questions[sheet] = []
        max_col = ws.max_column
        max_row = ws.max_row

        for col in range(1, max_col + 1):
            header = ws.cell(row=1, column=col).value
            if not header:
                continue

            q_text = str(header).strip()
            opts = []
            for row in range(2, max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is None or str(val).strip() == "":
                    continue
                opts.append(str(val).strip())

            questions[sheet].append({
                "text":    q_text,
                "options": opts
            })

    return questions


async def upload_langs(buffer: io.BytesIO):
    buffer.seek(0)
    wb = openpyxl.load_workbook(buffer, data_only=True)    
    ws = wb.active

    codes = [cell.value for cell in ws[1] if cell.value is not None]

    names = [cell.value for cell in ws[2] if cell.value is not None]

    lang = dict(zip(codes, names))

    translates = {}
    for row in ws.iter_rows(min_row=3, max_col=len(codes)+1, values_only=True):
        key = row[0]
        if not key:
            continue

        entry = {}
        for idx, code in enumerate(codes, start=1):
            val = row[idx]
            if val is not None and str(val).strip() != "":
                entry[code] = str(val).strip()
        translates[str(key).strip()] = entry

    return lang, translates


async def upload_prices(buffer: io.BytesIO):
    buffer.seek(0)
    wb = openpyxl.load_workbook(buffer, data_only=True)    
    ws = wb.active

    prices = {}

    for row in ws.iter_rows(max_row=ws.max_row, values_only=True):
        key = row[0]
        if not key:
            continue

        prices[key] = int(row[1])

    return prices